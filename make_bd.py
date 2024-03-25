from sqlalchemy import create_engine
from data.models import *
from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker
import sqlite3 as sql

engine = create_engine('sqlite:///test_transfer.db')
last_db_engine = sql.connect("db.sqlite3")

Model.metadata.create_all(engine)

Session = sessionmaker(bind=engine)
session = Session()

old_db = last_db_engine.cursor()

countries = {
    1: 1,
    2: 2,
    3: None,
    4: 3,
    5: 4,
    6: 5,
    7: 6,
    9: 7,
    10: None,
    11: 8,
    12: None,
    13: 9,
    14: 10,
    15: 11,
    16: 12,
    17: 13,
    18: 14,
    19: 15,
    20: 16,
    21: 17,
    22: 18,
    23: 19,
    24: 20
}

wallets = {
    47: 'Семён',
    4: 'Вова',
    46: 'Соня',
    34: 'Мактал',
    20: 'Егор',
    44: 'Борис',
    23: 'Энди',
    1: 'Слава',
    26: 'Соня'
}

# Загружаем данные из файла "Файл для БД.xmls"
workbook = load_workbook(filename="Файл для БД.xlsx")

# Получаем лист "СТРАНЫ И ЭМОДЗИ"
sheet_countries = workbook["СТРАНЫ И ЭМОДЗИ"]

# Создаем страны в базе данных
for row in sheet_countries.iter_rows(min_row=2, values_only=True):
    country_name = row[0]
    country_emoji = row[1]

    # Проверяем, существует ли уже страна с таким названием
    existing_country = session.query(Country).filter_by(
        name=country_name).first()
    if not existing_country:
        # Создаем новую страну, если она не существует
        country = Country(name=country_name, flag=country_emoji)
        session.add(country)

# Получаем лист "ШАБЛОНЫ БК"
sheet_templates = workbook["ШАБЛОНЫ БК"]

# Создаем шаблоны букмекеров в базе данных
for row in sheet_templates.iter_rows(min_row=2, values_only=True):
    country_name = row[0]
    bookmaker_name = str(row[1]).capitalize()
    bookmaker_percentage = row[2]

    # Получаем страну по названию
    country = session.query(Country).filter_by(name=country_name).first()
    if country:
        # Проверяем, существует ли уже шаблон с таким названием для данной страны
        existing_template = session.query(Template).filter_by(
            name=bookmaker_name, country=country).first()
        if not existing_template:
            # Создаем новый шаблон, если он не существует
            template = Template(name=bookmaker_name, country=country,
                                employee_percentage=bookmaker_percentage)
            session.add(template)

# Получаем лист "ПАРТНЕРЫ"
sheet_partners = workbook["ПАРТНЕРЫ"]

# Создаем партнеров (источники) в базе данных
for row in sheet_partners.iter_rows(min_row=2, max_col=1, values_only=True):
    partner_name = row[0]

    # Проверяем, существует ли уже партнер с таким названием
    existing_source = session.query(Source).filter_by(
        name=partner_name).first()
    if not existing_source:
        # Создаем новый источник, если он не существует
        source = Source(name=partner_name)
        session.add(source)

# Создаем пользователей в базе данных
for user in old_db.execute(
        "SELECT * FROM Users WHERE is_accepted=1").fetchall():
    user_id = user[1]
    username = user[2]
    first_name = user[3]

    exist = session.query(Employee).filter_by(
        id=user_id
    ).first()

    if not exist:
        employee = Employee(
            id=user_id,
            username=username,
            name=first_name
        )
        session.add(employee)


# Импорт букмейкеров
workbook = load_workbook(filename="БК и балансы.xlsx")
sheet = workbook.active


for row in sheet.iter_rows(min_row=2, values_only=True):
    country_name = row[0]
    bk_name = row[1].capitalize()
    profile_name = row[2].capitalize()

    balance_static = int(row[3])
    balance_active = int(row[3]) if isinstance(row[4], str) else int(row[4])

    country = session.query(Country).filter_by(
        name=country_name
    ).first()

    template = session.query(Template).filter_by(
        name=bk_name
    ).first()

    if country and template:
        exist = session.query(Bookmaker).filter_by(
            name=profile_name, template_id=template.id, country_id=country.id
        ).first()

        if not exist:
            bookmaker = Bookmaker(
                name=profile_name,
                template_id=template.id,
                country_id=country.id,
                bk_name=bk_name
            )
            session.add(bookmaker)

            bookmaker = session.query(Bookmaker).filter_by(
                name=profile_name, template_id=template.id,
                country_id=country.id, bk_name=bk_name
            ).first()

            transaction_deposit = Transaction(
                sender_wallet_id=None,
                receiver_wallet_id=None,
                sender_bookmaker_id=None,
                receiver_bookmaker_id=bookmaker.id,
                amount=balance_static,
                commission=None,
                from_=None,
                where="deposit"
            )
            session.add(transaction_deposit)

            transaction_balance = Transaction(
                sender_wallet_id=None,
                receiver_wallet_id=None,
                sender_bookmaker_id=None,
                receiver_bookmaker_id=bookmaker.id,
                amount=(balance_active - balance_static),
                commission=None,
                from_=None,
                where="balance"
            )
            session.add(transaction_balance)

# Импорт кошельков
for row in old_db.execute(f"SELECT * FROM Wallets").fetchall():
    if row[0] not in list(wallets.keys()):
        continue

    general_type = 'Карта' if row[1] == 'card' else 'Binance' 
    wallet_name = wallets[row[0]]
    wallet_type = 'Страна'
    country_id = countries[row[3]]
    deposit = row[5]

    country = session.query(Country).filter_by(
        id=country_id
    ).first()

    if country:
        wallet = Wallet(name=wallet_name, wallet_type=wallet_type,
                               general_wallet_type=general_type,
                               deposit=deposit, country_id=country.id)
        session.add(wallet)
    
    


# Сохраняем изменения в базе данных
session.commit()
